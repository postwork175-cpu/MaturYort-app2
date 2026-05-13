import streamlit as st
import io
import numpy as np
from PIL import Image, ImageOps, ImageEnhance, ImageDraw
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter

# --- КОНФИГУРАЦИЯ ---
PALETTE_HEX = {1: "FFFFFF", 2: "D3D3D3", 3: "808080", 4: "404040", 5: "000000"}
PALETTE_RGB = {k: tuple(int(v[i:i+2], 16) for i in (0, 2, 4)) for k, v in PALETTE_HEX.items()}
PALETTE_VALS = [PALETTE_RGB[i] for i in range(1, 6)]
MAX_DETAILS_PER_COLOR = 1100 
SMART_LIMITS = {i: MAX_DETAILS_PER_COLOR for i in range(1, 6)}

# --- ФУНКЦИИ ОБРАБОТКИ ---

def get_col_name(n):
    res = ""
    while n > 0:
        n, rem = divmod(n - 1, 26)
        res = chr(65 + rem) + res
    return res

def create_pdf(scheme_data, usage, fw, fh):
    pdf_buffer = io.BytesIO()
    c = canvas.Canvas(pdf_buffer, pagesize=A4)
    margin = 20 * mm
    cell_size = 5.5 * mm
    sectors = [(0, 0, 32, 32, "ЛЕВЫЙ ВЕРХ"), (32, 0, fw, 32, "ПРАВЫЙ ВЕРХ"), (0, 32, 32, fh, "ЛЕВЫЙ НИЗ"), (32, 32, fw, fh, "ПРАВЫЙ НИЗ")]
    for start_x, start_y, end_x, end_y, label in sectors:
        if start_x >= fw or start_y >= fh: continue
        c.setFont("Helvetica-Bold", 14)
        c.drawString(margin, 285 * mm, f"ИНСТРУКЦИЯ - СЕКТОР: {label}")
        for y in range(start_y, end_y):
            if y >= fh: break
            py = (270 * mm) - (y - start_y + 1) * cell_size
            for x in range(start_x, end_x):
                if x >= fw: break
                px = margin + (x - start_x) * cell_size
                c.setStrokeColorRGB(0.8, 0.8, 0.8)
                c.rect(px, py, cell_size, cell_size)
                val = scheme_data[y][x]
                c.setFillColorRGB(0, 0, 0) if val < 4 else c.setFillColorRGB(0.6, 0.6, 0.6)
                c.setFont("Helvetica-Bold", 10)
                c.drawCentredString(px + cell_size/2, py + 2.5 * mm, str(val))
        c.showPage()
    c.save()
    pdf_buffer.seek(0)
    return pdf_buffer

def create_excel(scheme_data, side_x, side_y):
    wb = openpyxl.Workbook()
    ws = wb.active
    for y in range(1, side_y + 1):
        ws.row_dimensions[y].height = 22
        for x in range(1, side_x + 1):
            if y == 1: ws.column_dimensions[get_column_letter(x)].width = 4
            val = scheme_data[y-1][x-1]
            cell = ws.cell(row=y, column=x, value=val)
            cell.fill = PatternFill(start_color=PALETTE_HEX[val], fill_type="solid")
            cell.font = Font(color="000000" if val < 4 else "FFFFFF", size=8)
            cell.alignment = Alignment(horizontal='center', vertical='center')
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

def create_preview(scheme_data, side_x, side_y):
    scale = 12
    img = Image.new('RGB', (side_x * scale, side_y * scale))
    pixels = img.load()
    for y in range(side_y):
        for x in range(side_x):
            color_rgb = PALETTE_RGB[scheme_data[y][x]]
            for i in range(scale):
                for j in range(scale):
                    pixels[x*scale+i, y*scale+j] = color_rgb
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)
    return buf

def process_mosaic_core(image_bytes, tw, th):
    img = Image.open(io.BytesIO(image_bytes)).convert('L')
    img = img.resize((tw, th), resample=Image.LANCZOS)
    img = ImageOps.autocontrast(img, cutoff=1)
    img = ImageEnhance.Sharpness(img).enhance(2.5)
    px = np.array(img, dtype=float)
    sch = np.zeros((th, tw), dtype=int)
    use = {i+1: 0 for i in range(5)}
    for y in range(th):
        for x in range(tw):
            dist_list = [(abs(px[y,x] - sum(c)/3), i+1, sum(c)/3) for i, c in enumerate(PALETTE_VALS)]
            dist_list.sort()
            idx, nval = dist_list[0][1], dist_list[0][2]
            for d, i, v in dist_list:
                if use[i] < SMART_LIMITS[i]:
                    idx, nval = i, v
                    break
            sch[y,x], use[idx], err = idx, use[idx]+1, px[y,x]-nval
            if x+1<tw: px[y,x+1] += err*7/16
            if y+1<th:
                if x>0: px[y+1,x-1] += err*3/16
                px[y+1,x] += err*5/16
                if x+1<tw: px[y+1,x+1] += err*1/16
    return sch.tolist(), use, tw, th

# --- ИНТЕРФЕЙС САЙТА ---
st.set_page_config(page_title="MaturYort", layout="centered")
st.title("🖼 Мозаика MaturYort")

uploaded_file = st.file_uploader("Загрузите фото", type=["jpg", "png", "jpeg"])

if uploaded_file:
    img_content = uploaded_file.read()
    st.image(img_content, width=300)
    
    if st.button("Создать схему"):
        # Открываем изображение для получения его реальных размеров
        im = Image.open(io.BytesIO(img_content))
        w, h = im.size
        
        # Вычисляем коэффициент масштабирования относительно максимальной стороны в 64 пикселя
        r = 64 / max(w, h)
        tw, th = int(w * r), int(h * r)
        with st.spinner("Генерирую..."):
            sch, use, fw, fh = process_mosaic_core(img_content, tw, th)
            
            st.image(create_preview(sch, fw, fh), caption="Превью")
            
            col1, col2 = st.columns(2)
            col1.download_button("📄 Скачать PDF", create_pdf(sch, use, fw, fh), "instruction.pdf")
            col2.download_button("📊 Скачать Excel", create_excel(sch, fw, fh), "scheme.xlsx")
            
            for k, v in use.items():
                st.write(f"Цвет {k}: {v} шт.")  
